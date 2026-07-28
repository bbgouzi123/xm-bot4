import os
from typing import List, Dict, Optional
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelParser:
    """Excel 客户资料与营销批号解析器 (Phase 9 升级)"""

    @staticmethod
    def parse_friend_list(file_path: str) -> List[Dict]:
        """解析将待添加的拓客资源 Excel 文件
        
        支持的表头识别匹配，哪怕用户排版乱了也能自动找：
        - 手机号/微信号/wxid
        - 备注名称/姓名/客户名
        - 标签 (可采用逗号隔开)
        """
        if not HAS_OPENPYXL:
            logger.error('openpyxl 未安装，无法解析 Excel')
            raise RuntimeError("请通过 pip install openpyxl 补齐依赖")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f'文件不存在: {file_path}')

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 第一行读取表头并归一化索引
            headers = [str(cell).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            wx_col = -1
            name_col = -1
            tag_col = -1
            
            for idx, h in enumerate(headers):
                if h in ('微信号', '手机号', 'wxid', '联系方式', '账号'):
                    wx_col = idx
                elif h in ('备注名称', '姓名', '客户', '昵称', '客户名'):
                    name_col = idx
                elif h in ('标签', '群体', 'tags'):
                    tag_col = idx
                    
            # 防呆：如果没写表头，默认猜它 0 是微信号，1 是备注，2 是标签
            if wx_col == -1: wx_col = 0
            if name_col == -1: name_col = 1
            if tag_col == -1: tag_col = 2
            
            friends = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= wx_col or not row[wx_col]:
                    continue
                    
                val_wxid = str(row[wx_col]).strip()
                if not val_wxid:
                    continue
                    
                val_name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ''
                val_tag = str(row[tag_col]).strip() if len(row) > tag_col and row[tag_col] else ''
                
                # 号码清洗：去掉可能带来的异常字符等
                val_wxid = val_wxid.replace(" ", "").replace("-", "")
                
                friend = {
                    'wxid': val_wxid,
                    'remark': val_name,
                    'tags': val_tag,
                }
                friends.append(friend)

            wb.close()
            logger.info(f'[Excel 分析仪] 成功清洗提取: {len(friends)} 个拓客目标')
            return friends

        except Exception as e:
            logger.error(f'[Excel 分析仪] 读取报错: {e}')
            raise 

    @staticmethod
    def parse_mass_send_list(file_path: str) -> List[str]:
        """解析群发目标独立白名单"""
        if not HAS_OPENPYXL:
            return []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            targets = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    targets.append(str(row[0]).strip())

            wb.close()
            return targets
        except Exception as e:
            logger.error(f'[Excel 分析仪] 群发名单读取失败: {e}')
            return []
