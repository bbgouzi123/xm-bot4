"""
微信联系人与群聊数据导出服务（生成 Excel 或 CSV 资产）
"""
import os
import tempfile
import logging
from fastapi.responses import FileResponse

logger = logging.getLogger(__name__)

def get_visible_len(val):
    if val is None:
        return 0
    s = str(val)
    length = 0
    for char in s:
        if ord(char) > 127:
            length += 2  # 中文字符宽度算 2
        else:
            length += 1  # 英文/数字算 1
    return length

def beautify_sheet(ws, has_image_support, accounts_dir, items_list):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except Exception:
        has_image_support = False

    # 样式设计
    fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="0F172A")
    font_data = Font(name="Microsoft YaHei", size=10, color="334155")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # 表头行高与样式
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    # 数据行高、内容和样式
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42
        item = items_list[row_idx - 2]
        wxid = item.get("wxid") or ""
        
        # 尝试在该行 A 列插入头像图片
        avatar_path = os.path.join(accounts_dir, f"{wxid}.png") if wxid else ""
        if has_image_support and avatar_path and os.path.exists(avatar_path):
            try:
                img = OpenpyxlImage(avatar_path)
                img.width = 56
                img.height = 56
                ws.add_image(img, f"A{row_idx}")
            except Exception as e_img:
                logger.debug(f"Failed to insert excel avatar image for {wxid}: {e_img}")
                ws.cell(row=row_idx, column=1, value="")
        else:
            ws.cell(row=row_idx, column=1, value="")

        # 为所有单元格应用基本样式与网格线
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if col_idx in (1, 2, 4):  # 头像、序号、群成员数/托管状态等简短字段居中排版
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # 自适应列宽
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            ws.column_dimensions[col_letter].width = 7.2  # 头像列固定宽度，契合 56px 图片大小
            continue
        max_len = 0
        for cell in col:
            # 忽略表头和空单元格的非理性计算
            max_len = max(max_len, get_visible_len(cell.value))
        # 加上内边距
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11)


def do_export_contacts(export_type: str = "friend", is_desktop: bool = False, selected_ids: str = ""):
    from src.utils.contacts_cache import contacts_cache
    from src.crm.account_data import get_active_account, ACCOUNTS_DIR
    
    active_id = get_active_account()
    selected_set = set(selected_ids.split(",")) if selected_ids else None
    
    if export_type == "group":
        groups = contacts_cache.get_groups(active_id)
        if not groups:
            try:
                contacts_cache.load_from_cloud()
                groups = contacts_cache.get_groups(active_id)
            except Exception:
                pass
        if not groups:
            groups = []
        
        # 去重
        dedup_map = {}
        for g in groups:
            name = g.get("name")
            if name:
                dedup_map[name] = g
            else:
                dedup_map[id(g)] = g
        groups_list = list(dedup_map.values())
        if selected_set:
            groups_list = [g for g in groups_list if g.get("wxid") in selected_set or g.get("name") in selected_set]
        
        filename = f"wechat_groups_{active_id}.xlsx"
        if is_desktop:
            from pathlib import Path
            downloads_dir = Path.home() / "Downloads"
            downloads_dir.mkdir(parents=True, exist_ok=True)
            file_path = os.path.join(downloads_dir, filename)
            counter = 1
            name_parts = os.path.splitext(filename)
            while os.path.exists(file_path):
                file_path = os.path.join(downloads_dir, f"{name_parts[0]}_{counter}{name_parts[1]}")
                counter += 1
            filename = os.path.basename(file_path)
        else:
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, filename)
        
        has_xlsx = False
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "群聊列表"
            headers = ["头像", "序号", "群名/群昵称", "微信内部ID (wxid)", "群成员数", "同步时间"]
            ws.append(headers)
            for idx, g in enumerate(groups_list, 1):
                ws.append([
                    "",  # 留给头像插图位置
                    idx,
                    g.get("name") or "",
                    g.get("wxid") or "",
                    len(g.get("members") or []),
                    g.get("syncTime") or ""
                ])
            
            beautify_sheet(ws, True, ACCOUNTS_DIR, groups_list)
            wb.save(file_path)
            has_xlsx = True
        except ImportError:
            import csv
            filename = f"wechat_groups_{active_id}.csv"
            if is_desktop:
                from pathlib import Path
                downloads_dir = Path.home() / "Downloads"
                downloads_dir.mkdir(parents=True, exist_ok=True)
                file_path = os.path.join(downloads_dir, filename)
                counter = 1
                name_parts = os.path.splitext(filename)
                while os.path.exists(file_path):
                    file_path = os.path.join(downloads_dir, f"{name_parts[0]}_{counter}{name_parts[1]}")
                    counter += 1
                filename = os.path.basename(file_path)
            else:
                temp_dir = tempfile.gettempdir()
                file_path = os.path.join(temp_dir, filename)
            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["头像", "序号", "群名/群昵称", "微信内部ID (wxid)", "群成员数", "同步时间"])
                for idx, g in enumerate(groups_list, 1):
                    writer.writerow([
                        "",
                        idx,
                        g.get("name") or "",
                        g.get("wxid") or "",
                        len(g.get("members") or []),
                        g.get("syncTime") or ""
                    ])
                    
        if is_desktop:
            from src.utils.response import ok
            return ok({
                "path": str(file_path),
                "filename": filename
            })

        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if has_xlsx else "text/csv"
        )
        
    else:
        friends = contacts_cache.get_friends(active_id)
        if not friends:
            try:
                contacts_cache.load_from_cloud()
                friends = contacts_cache.get_friends(active_id)
            except Exception:
                pass
        if not friends:
            friends = []
            
        sys_prefixes = ("新的朋友", "公众号", "企业微信联系人", "群聊", "标签", "服务号", "我的企业", "联系人", "文件传输助手")
        dedup_map = {}
        for f in friends:
            name = f.get("name", "").strip()
            is_sys = False
            if len(name) == 1 and name.isascii() and name.isalpha():
                is_sys = True
            for pre in sys_prefixes:
                if name.startswith(pre):
                    suffix = name[len(pre):].strip()
                    if not suffix or suffix.isdigit():
                        is_sys = True
                        break
                elif name == pre:
                    is_sys = True
                    break
            if not is_sys:
                dedup_key = f.get("wxid")
                if not dedup_key:
                    dedup_key = name
                if dedup_key:
                    if dedup_key not in dedup_map:
                        dedup_map[dedup_key] = f
                else:
                    dedup_map[id(f)] = f
                    
        friends_list = list(dedup_map.values())
        if selected_set:
            friends_list = [f for f in friends_list if f.get("wxid") in selected_set or f.get("name") in selected_set or f.get("nickname") in selected_set]
        
        filename = f"wechat_friends_{active_id}.xlsx"
        if is_desktop:
            from pathlib import Path
            downloads_dir = Path.home() / "Downloads"
            downloads_dir.mkdir(parents=True, exist_ok=True)
            file_path = os.path.join(downloads_dir, filename)
            counter = 1
            name_parts = os.path.splitext(filename)
            while os.path.exists(file_path):
                file_path = os.path.join(downloads_dir, f"{name_parts[0]}_{counter}{name_parts[1]}")
                counter += 1
            filename = os.path.basename(file_path)
        else:
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, filename)
        
        has_xlsx = False
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "好友列表"
            headers = ["头像", "序号", "微信昵称", "备注名称", "微信号 (alias)", "微信原始ID (wxid)", "地区", "添加来源", "个性签名", "标签/Tag", "是否托管中"]
            ws.append(headers)
            for idx, f in enumerate(friends_list, 1):
                ws.append([
                    "",  # 留给头像插图位置
                    idx,
                    f.get("nickname") or f.get("name") or "",
                    f.get("remark") or "",
                    f.get("alias") or "",
                    f.get("wxid") or "",
                    f.get("region") or "",
                    f.get("source") or "",
                    f.get("signature") or "",
                    f.get("tag") or "",
                    "是" if f.get("is_takeover") else "否"
                ])
            
            beautify_sheet(ws, True, ACCOUNTS_DIR, friends_list)
            wb.save(file_path)
            has_xlsx = True
        except ImportError:
            import csv
            filename = f"wechat_friends_{active_id}.csv"
            if is_desktop:
                from pathlib import Path
                downloads_dir = Path.home() / "Downloads"
                downloads_dir.mkdir(parents=True, exist_ok=True)
                file_path = os.path.join(downloads_dir, filename)
                counter = 1
                name_parts = os.path.splitext(filename)
                while os.path.exists(file_path):
                    file_path = os.path.join(downloads_dir, f"{name_parts[0]}_{counter}{name_parts[1]}")
                    counter += 1
                filename = os.path.basename(file_path)
            else:
                temp_dir = tempfile.gettempdir()
                file_path = os.path.join(temp_dir, filename)
            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["头像", "序号", "微信昵称", "备注名称", "微信号 (alias)", "微信原始ID (wxid)", "地区", "添加来源", "个性签名", "标签/Tag", "是否托管中"])
                for idx, f_item in enumerate(friends_list, 1):
                    writer.writerow([
                        "",
                        idx,
                        f_item.get("nickname") or f_item.get("name") or "",
                        f_item.get("remark") or "",
                        f_item.get("alias") or "",
                        f_item.get("wxid") or "",
                        f_item.get("region") or "",
                        f_item.get("source") or "",
                        f_item.get("signature") or "",
                        f_item.get("tag") or "",
                        "是" if f_item.get("is_takeover") else "否"
                    ])
                    
        if is_desktop:
            from src.utils.response import ok
            return ok({
                "path": str(file_path),
                "filename": filename
            })

        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if has_xlsx else "text/csv"
        )


